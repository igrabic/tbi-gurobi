"""
TBI - Data generation for the optimization model
Version: 2.0
Authors: Antonio Karneluti, Ivan Grabić, Filip Rukavina, Marko Kovačević, Mario Vašak
University of Zagreb, Faculty of Electrical Engineering and Computing
Department for Control and Computer Engineering
Laboratory for Renewable Energy Systems (LARES)
--------------------------------------------------------
This file collects all the data needed for the optimization model and generates the data structure.
"""

import numpy as np
import datetime
import pandas as pd
import json
import requests
from io import StringIO
import re
from scipy import interpolate
import openpyxl


class Structure:
    pass


# TODO remove d
def generate_data(tbi="TBI1", scenario="Scenario 1"):
    """
    Prepares data needed for the LP: transforms parameters, resamples profiles, etc.

    Arguments:

    Output:
        p: Structure, generated data
    """

    # initialize data structure which keeps all the generated data
    p = Structure()

    p = get_data_from_excel(tbi, scenario)

    ############################################################################

    p.start_date = datetime.date(2025, 1, 1)  # start date of simulation [dd.mm.yyyy]
    p.end_date = datetime.date(2025, 12, 31)  # last date of simulation [dd.mm.yyyy]

    p.N_months = 12
    p.days_in_month = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
    p.N_days = int(
        (p.end_date - p.start_date).days + 1
    )  # number of days in simulation - one year [-]
    p.el_N_s = int(24 / p.el_t_s) * p.N_days  # number of time steps
    p.el_t_15min = 0.25  # time step for 15 min [h]
    p.N_s_day = int(24 / p.el_t_s)

    ## PV panels
    # for each array, get the power profile from PVGIS
    # initialize the array of PV panels
    p.PV_P_ref_full = np.zeros(
        (p.PV_N_arr, 35040), dtype=np.float32
    )  # power of PV panel array [kW]
    p.PV_P_ref = np.zeros(
        (p.PV_N_arr, p.el_N_s), dtype=np.float32
    )  # power of PV panel array [kW]
    p.PV_P_ref_peak = np.zeros(
        p.PV_N_arr, dtype=np.float32
    )  # peak power of PV panel array [kW]
    for i in range(p.PV_N_arr):
        p.PV_P_ref_full[i] = (
            get_pvgis_power_array(
                lat=p.latitude, lon=p.longitude, azimuth=0, tilt=45, year=2019
            )
            / 1000
        )
        p.PV_P_ref[i] = np.mean(
            p.PV_P_ref_full[i].reshape(-1, int(len(p.PV_P_ref_full[i]) / p.el_N_s)),
            axis=1,
        )
        p.PV_P_ref_peak[i] = max(p.PV_P_ref[i])  # peak power of PV panel array [kW]

    p.PV_P_exist = np.zeros(
        (p.el_N_s, p.PV_N_arr)
    )  # existing power of PV panel array [kW]
    p.PV_P_exist_peak = np.zeros(p.PV_N_arr)  # peak power of PV panel array [kW]

    ## Battery
    p.batt_E_exist = 0  # existing battery energy [kWh]
    p.batt_c_deg = p.batt_c_rpl / (2 * p.batt_N_cyc * p.batt_s_DoD)

    # Power converter
    p.PC_P_exist = 0  # existing power converter [kW]

    # Battery energy storage system (BESS)
    p.BESS_t_s = p.el_t_s
    p.BESS_E_0_min = p.batt_E_new_min  # minimum energy in the battery [kWh]
    p.BESS_E_0_max = p.batt_E_new_max  # maximum energy in the battery [kWh]

    ## Renewable energy plants (REP)
    p.N_REP = 0  # number of REP plants [-]
    p.REP_P_ref = np.zeros(p.el_N_s)  # power of REP plants [kW]

    # Electricity grid
    p = get_el_price(
        p, day_ahead_prices=True
    )  # price of electricity from and to the grid [€/kWh] - EGC_c_en_sell, EGC_c_en_buy
    p.EGC_P_fix = get_consumption_profile()  # fixed consumption profile [kW]
    p.EGC_P_fix_peak = get_consumption_peak(
        p.EGC_P_fix
    )  # fixed profile peak power [€/kW]
    p.EGC_P_fix = np.mean(
        p.EGC_P_fix.reshape(-1, int(len(p.EGC_P_fix) / p.el_N_s)), axis=1
    )

    p.EGC_s_BPP_min = 0.7
    p.EGC_s_BPP_max = 1.3
    p.EGC_alpha_BPP_add = 1.5

    p.C_noinv_remain = np.sum(p.EGC_P_fix * p.EGC_c_en_buy) + np.sum(
        p.EGC_c_peak * p.EGC_P_fix_peak
    )  # [€]

    return p


def get_consumption_peak(consumption):
    """
    Get peak power of the consumption profile [kW]
    """
    days_in_month = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
    samples_per_day = int(24 / 0.25)  # assuming el_t_s is 0.25 hours (15 minutes)
    existing_peaks = np.zeros((12), dtype=np.float32)
    start_idx = 0
    for month_idx, days in enumerate(days_in_month):
        end_idx = start_idx + days * samples_per_day
        existing_peaks[month_idx] = max(consumption[start_idx:end_idx])
        start_idx = end_idx
    return existing_peaks


def get_data_from_excel(tbi, scenario):
    """
    Read data from the Excel file and fill the data structure with the values.
    Read the names of the columns from the first row of the Excel file.
    The first column is the name of the parameter, the column_name parameter is the column with the value.
    """

    # Use the structure
    p = Structure()

    # Read the Excel file to get the parameters using openpyxl
    excel_path = "./parameters.xlsx"
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
    except FileNotFoundError:
        raise FileNotFoundError(
            f"Excel file '{excel_path}' not found. Please ensure the file exists."
        )
    if tbi not in wb.sheetnames:
        raise ValueError(
            f"Sheet '{tbi}' not found in the Excel file. Available sheets: {wb.sheetnames}"
        )
    # Read the specific sheet
    ws = wb[tbi]
    # Convert the sheet to a DataFrame
    data = ws.values
    columns = next(data)  # Get the first row as column names
    df = pd.DataFrame(data, columns=columns)

    # The first column contains parameter names, the rest are scenarios
    if scenario not in df.columns:
        raise ValueError(f"Scenario '{scenario}' not found in sheet '{tbi}'.")

    param_names = df.iloc[:, 0]
    values = df[scenario]

    for param, val in zip(param_names, values):
        if isinstance(val, str) and (";" in val or "," in val):
            # Try to parse as array (semicolon or comma separated)
            sep = ";" if ";" in val else ","
            arr = [
                float(x.replace(",", ".").strip())
                for x in val.split(sep)
                if x.strip() != ""
            ]
            val = np.array(arr, dtype=np.float32)
        elif isinstance(val, str):
            # Try to parse as float (replace comma with dot)
            try:
                val = float(val.replace(",", "."))
            except Exception:
                pass
        elif isinstance(val, (int, float, np.integer, np.floating)):
            pass
        # Set attribute in structure
        setattr(p, str(param), val)
        # print(f"Set parameter {param} to {val}")

    return p


def get_el_price(p, day_ahead_prices=False):
    if not day_ahead_prices:
        p.EGC_c_en_buy = np.array([], dtype=np.float32)
        p.EGC_c_en_sell = np.array([], dtype=np.float32)
        p_el_weekday_consume = np.concatenate(
            (
                np.tile(p.p_el_n_consume, int(p.el_day_start / p.el_t_15min)),
                np.tile(
                    p.p_el_d_consume,
                    int((p.el_day_end - p.el_day_start) / p.el_t_15min),
                ),
                np.tile(p.p_el_n_consume, int((24 - p.el_day_end) / p.el_t_15min)),
            )
        )
        p_el_weekday_supply = np.concatenate(
            (
                np.tile(p.p_el_n_supply, int(p.el_day_start / p.el_t_15min)),
                np.tile(
                    p.p_el_d_supply, int((p.el_day_end - p.el_day_start) / p.el_t_15min)
                ),
                np.tile(p.p_el_n_supply, int((24 - p.el_day_end) / p.el_t_15min)),
            )
        )
        for day in range(p.start_date.toordinal(), p.end_date.toordinal() + 1):
            p.EGC_c_en_buy = np.concatenate((p.EGC_c_en_buy, p_el_weekday_consume))
            p.EGC_c_en_sell = np.concatenate((p.EGC_c_en_sell, p_el_weekday_supply))
        p.EGC_c_en_sell = np.mean(
            p.EGC_c_en_sell.reshape(-1, int(len(p.EGC_c_en_sell) / p.el_N_s)), axis=1
        )
        p.EGC_c_en_buy = np.mean(
            p.EGC_c_en_buy.reshape(-1, int(len(p.EGC_c_en_buy) / p.el_N_s)), axis=1
        )
    else:
        p.EGC_c_en_buy = get_day_ahead_prices(p) + 1e-3
        p.EGC_c_en_sell = np.where(
            p.EGC_c_en_buy >= 0, p.EGC_c_en_buy * 0.9, p.EGC_c_en_buy * 1.1
        )

    return p


def get_day_ahead_prices(p):
    """
    Reads the day-ahead prices from a CSV file and returns it as a DataFrame.
    """
    df = pd.read_csv(
        "./day-ahead_prices_CRO_2024.csv",
        skiprows=1,
        sep=";",
        names=["timestamp", "price"],
    )
    df["timestamp"] = pd.to_datetime(df["timestamp"], format="%d.%m.%Y %H:%M")
    df.set_index("timestamp", inplace=True)
    # Ensure the index is a DatetimeIndex with a fixed frequency before resampling
    df = df.resample("{}h".format(p.el_t_s)).mean().interpolate()
    df["price"] = df["price"].ffill() / 1000  # Fill any remaining NaNs if needed
    prices = df.price.values[: p.el_t_s * 24 * 365]  # Return only the first year

    assert (
        len(prices) == p.el_N_s
    ), f"Expected {p.el_N_s} prices, but got {len(prices)}."
    if np.any(np.isnan(prices)) or np.any(np.isinf(prices)):
        raise ValueError("Day-ahead prices contain NaN or inf values.")

    return prices


def get_pvgis_power_array(lat, lon, azimuth, tilt, year=2020, peakpower=1.0, loss=0):
    """
    Fetches hourly PV power output and interpolates to 15-minute resolution,
    then pads with zeros to ensure exactly 35,040 samples per year.

    Returns:
        np.ndarray: Array of power values in watts (W) at 15-minute intervals (35,040 values).
    """
    url = "https://re.jrc.ec.europa.eu/api/v5_2/seriescalc"
    params = {
        "lat": lat,
        "lon": lon,
        "startyear": year,
        "endyear": year,
        "pvcalculation": 1,
        "peakpower": peakpower,
        "loss": loss,
        "angle": tilt,
        "aspect": azimuth,
        "outputformat": "csv",
        "usehorizon": 1,
        "optimalangles": 0,
        "raddatabase": "PVGIS-ERA5",
        "components": 1,
        "hourlyvalues": 1,
    }

    response = requests.get(url, params=params)
    if response.status_code != 200:
        raise Exception(
            f"Error fetching data: {response.status_code} - {response.text}"
        )
    content = response.text
    lines = content.splitlines()
    valid_lines = [line for line in lines if re.match(r"^\d{8}:\d{4}|^time", line)]
    cleaned_csv = "\n".join(valid_lines)

    df = pd.read_csv(StringIO(cleaned_csv), sep=",")
    df["time"] = pd.to_datetime(df["time"], format="%Y%m%d:%H%M", utc=True)
    if "P" not in df.columns:
        raise Exception("Power column 'P' not found in the data.")

    hourly_series = df.set_index("time")["P"]
    full_year_index = pd.date_range(
        start=f"{year}-01-01 00:00:00",
        end=f"{year}-12-31 23:45:00",
        freq="15min",
        tz="UTC",
    )
    f = interpolate.interp1d(
        hourly_series.index.astype(np.int64),
        hourly_series.values,
        kind="linear",
        fill_value=0,
        bounds_error=False,
    )

    interpolated_values = f(full_year_index.astype(np.int64))
    expected_length = 35040
    if len(interpolated_values) < expected_length:
        zeros_to_add = expected_length - len(interpolated_values)
        interpolated_values = np.append(interpolated_values, np.zeros(zeros_to_add))
    elif len(interpolated_values) > expected_length:
        interpolated_values = interpolated_values[:expected_length]
    interpolated_values = np.clip(interpolated_values, 0, None)

    return interpolated_values


def get_consumption_profile():
    """
    Read csv from data folder and return the consumption power profile
    """
    df = pd.read_csv(
        "./metalproduct_cons_at_EGC_2024.csv",
        usecols=[1],
        skiprows=1,
        header=None,
        names=["P"],
    )
    df["timestamp"] = pd.date_range(
        start="2024-01-01 00:00:00", end="2024-12-31 23:45:00", freq="15min"
    )
    df = df[~((df["timestamp"].dt.month == 12) & (df["timestamp"].dt.day == 31))]
    df = df["P"].values.flatten()
    return df


def get_no_inv_peak(p):
    """
    Get peak power of the existing power capacity [kW]
    """
    days_in_month = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
    samples_per_day = int(24 / p.el_t_s)
    existing_peaks = np.zeros((12), dtype=np.float32)
    start_idx = 0
    for month_idx, days in enumerate(days_in_month):
        end_idx = start_idx + days * samples_per_day
        existing_peaks[month_idx] = max(p.EGC_P_fix[start_idx:end_idx])
        start_idx = end_idx
    return existing_peaks


def get_el_prices(
    option,
    c_en_HT=0.147025352,
    c_en_LT=0.085168778,
    c_RES=0.013239,
    c_extra=0.0005,
    c_grid_HT=0.020811,
    c_grid_LT=0.010404,
    percent_sale=0.9,
    time_HT=6,
    time_LT=20,
    consumption=np.zeros(35040),
):
    timestamp = pd.date_range(
        start="2024-01-01 00:00",
        end="2024-12-30 23:45",
        freq="15min",
        tz="Europe/Zagreb",
    )
    timestamp_utc = timestamp.tz_convert("utc")
    c_buy = pd.Series(data=np.nan, index=timestamp)
    c_sell = pd.Series(data=np.nan, index=timestamp)
    cons = pd.Series(data=consumption, index=timestamp, name="P_dem")

    if option == "standard":

        # prices for buying electricity
        c_buy_HT = c_en_HT + c_RES + c_extra + c_grid_HT
        c_buy_LT = c_en_LT + c_RES + c_extra + c_grid_LT
        c_buy[(timestamp_utc.hour >= time_HT) & (timestamp_utc.hour < time_LT)] = (
            c_buy_HT
        )
        c_buy[(timestamp_utc.hour < time_HT) | (timestamp_utc.hour >= time_LT)] = (
            c_buy_LT
        )

        # prices for selling electricity
        # c_sell = np.zeros(12)
        # for month in range(1, 13):
        #     sum_HT =
        #
        #
        #     temp_cons = cons[timestamp.month == month]
        #
        #     weight_HT = cons[timestamp.mo]

    elif option == "enna":
        pass
